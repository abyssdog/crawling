import datetime
import os

from openpyxl import load_workbook
import pymysql.connections


def table_check(table):
    try:
        conn = pymysql.connect(host='127.0.0.1', user='root', port=3307, password='dangam1234', db='crawl',
                               charset='utf8mb4')
        curs = conn.cursor()
        sql = """SELECT * FROM {}""".format(table)
        curs.execute(sql)
        return True
    except pymysql.connections.err.ProgrammingError as e:
        print(e)
        return False
        # 만약 에러가 발생하면 해당 테이블이 없다는 것이므로
        # 테이블 생성.
        # 만약 에러가 아니면 해당 테이블로 업데이트 하기.


# 엑셀파일의 컬럼갯수에 따라 %s, row[0] 갯수 조절
def insert_excel_to_db():
    search_list = ["반려동물", "반려동물이벤트", "반려동물행사", "반려동물교육", "반려동물자격증", "반려동물직업",
                   "반려동물산업", "반려동물상품"]
    now = datetime.datetime.now()
    date_search = datetime.datetime.strftime(now, "%Y%m%d")
    date_now = datetime.datetime.strftime(now, "%Y-%m-%d")
    today_year = str(datetime.datetime.today().year)
    today_month = str(datetime.datetime.today().month)
    for key in search_list:
        # file_name = '{date}_blog_{key}'.format(date=date_search, key=key)
        file_name = '20200412_blog_{key}'.format(key=key)
        conn = pymysql.connect(host='127.0.0.1', user='root', port=3306, password='dangam1234', db='convention', charset='utf8mb4')
        try:
            with conn.cursor() as curs:
                presql = ''
                subsql = ''
                wb = load_workbook('{}.csv'.format(file_name), data_only=True)
                ws = wb['{}'.format(file_name)]
                max_row = ws.max_row  # 전체 row
                max_column = ws.max_column  # 전체 col
                templist = [[0 for col in range(max_column+1)] for row in range(max_row+1)]
                # file_name == table_name => 통일해서 알고리즘 진행되도록
                # presql = 'insert into pet_population ('
                presql = 'insert into naver_api ('
                for col in range(1, max_column+1):
                    if col == max_column:
                        presql += '{})'.format(ws.cell(row=1, column=col).value)
                    else:
                        presql += '{}, '.format(ws.cell(row=1, column=col).value)
                subsql = presql + ' value('
                # 병합된 셀이 있으면 체크해서 값 채우는 소스
                cnt = 0
                for row in range(2, ws.max_row+1):
                    for col in range(1, max_column+1):
                        # cell = ''
                        if ws.cell(row=row, column=col).value:
                            cell = ws.cell(row=row, column=col).value
                        else:
                            if col == 1:
                                cell = "한국"
                            elif 1 < col < 6:
                                cell = templist[row-1][col]
                            elif col == 7:
                                cell = "agency"
                            else:
                                # 이건 반려인 국외 컬럼용
                                # if col == 2 or col == 3 or col == 7 or col == 11 or col == 15 or col == 19 or col == 23 or col == 27 or col == 31 or col == 35 or col == 39:
                                # 이건 반려인 국내 컬럼용
                                # if col == 7 or col == 11 or col == 15 or col == 19 or col == 23 or col == 27 or col == 31 or col == 35 or col == 39:
                                if col == 8 or col == 12 or col == 16 or col == 20 or col == 24 or col == 28 or col == 32 or col == 36 or col == 40:
                                    cell = 0
                                else:
                                    cell = '-'
                        if col == max_column:
                            subsql += '"{}")'.format(cell)
                            print(subsql)  # 여기에 프린트대신 curs.execute(subsql)
                            #curs.execute(subsql)
                            subsql = presql + ' value('
                        else:
                            # true = int 만 해당, false = string
                            # 반려인 현황 - 국외 sheet 1,2
                            # if col == 2 or col == 3 or col == 7 or col == 11 or col == 15 or col == 19 or col == 23 or col == 27 or col == 31 or col == 35 or col == 39:
                            # 반려인 현황 - 국내 sheet 3,4
                            # if col == 7 or col == 11 or col == 15 or col == 19 or col == 23 or col == 27 or col == 31 or col == 35 or col == 39:
                            if col == 3 or col == 8 or col == 12 or col == 16 or col == 20 or col == 24 or col == 28 or col == 32 or col == 36 or col == 40:
                                subsql += '{}, '.format(cell)
                            else:
                                subsql += '"{}", '.format(cell)
                            templist[row][col] = cell
                #conn.commit()
        finally:
            conn.close()
            wb.close()


if __name__ == '__main__':
    # 일단은 수동으로 엑셀 입력, 나중에 프론트에서 엑셀파일 받아서 자동으로 입력.
    # insert_excel_to_db()
    # table_check("aaa")
    print(os.path.relpath('./originalDatas'))


