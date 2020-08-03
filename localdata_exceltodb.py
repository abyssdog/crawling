import datetime

import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook


# 엑셀파일 DB Insert
def get_last_data_key(key):
    conn = pymysql.connect(host='211.49.126.238', user='banji', port=3306,
                           password='banji2020@!', db='banji', charset='utf8mb4')
    curs = conn.cursor()
    sql = """ SELECT IFNULL(MAX(data_key)+1,1) AS data_key 
                FROM td_ods_{};
    """.format(key)
    curs.execute(sql)
    rows = curs.fetchall()
    return rows[0][0]


def get_last_meta_key(meta_key):
    conn = pymysql.connect(host='211.49.126.238', user='banji', port=3306,
                           password='banji2020@!', db='banji', charset='utf8mb4')
    curs = conn.cursor()
    '''sql = """ 
            SELECT IFNULL(MAX(collect_key)+1,1) AS data_key 
              FROM tm_data_collect 
             WHERE meta_key = '{}';
        """.format(meta_key)'''
    sql = """ 
            SELECT IFNULL(MAX(collect_key)+1,1) AS data_key 
              FROM tm_data_collect;
        """
    curs.execute(sql)
    rows = curs.fetchall()
    return rows[0][0]


def insert_excel_to_db(id, meta_key):
    now = datetime.datetime.now()
    create_date = datetime.datetime.strftime(now, "%Y%m%d%H%M%S")
    date_now = datetime.datetime.strftime(now, "%Y%m%d")
    conn = pymysql.connect(host='211.49.126.238', user='banji', port=3306,
                           password='banji2020@!', db='banji', charset='utf8mb4')
    last_collect_key = get_last_meta_key(meta_key)
    last_data_key = get_last_data_key(meta_key)
    data_count = last_data_key
    try:
        with conn.cursor() as curs:
            sql_data_collect = """
                insert into tm_data_collect values(
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s
                )
            """
            sql_ods = """
                insert into td_ods_{} values(%s, %s, %s,  
                    %s, %s, %s
            )""".format(meta_key)
            wb = load_workbook(r'C:\Users\work\Downloads\메타데이터 ver 0.2\{}.xlsx'.format(id), data_only=True)
            ws = wb['Sheet']

            iter_rows = iter(ws.rows)
            next(iter_rows)
            cnt = 0
            for row in iter_rows:
                if cnt == 0:
                    a = row[18].value.replace('-', '')
                    b = datetime.datetime.strptime(a, "%Y%m%d")
                    curs.execute(sql_data_collect, (
                        last_collect_key,  # collect_key
                        meta_key,  # meta_key
                        date_now,  # collect_sdate
                        date_now,  # collect_edate
                        create_date,  # create_date
                        row[15].value,  # create_user
                        '',  # collect_name
                        row[9].value,  # collect_type
                        row[10].value,  # collect_protocol
                        'N',  # flag_manual
                        row[6].value,  # source
                        '',  # source_url
                        '',  # manual_data_filename
                        '',  # manual_data_filepath
                        # datetime.datetime.strftime(row[18].value, "%Y%m%d%H%M%S"),  # update_date
                        datetime.datetime.strftime(b, "%Y%m%d"),  # update_date
                        row[16].value,  # updte_user
                        'Y'  # flag_active
                    ))
                    cnt += 1
                curs.execute(sql_ods, (  # t = row[19], url = 11 key = 19
                    data_count,  # content = 3, source = 6
                    last_collect_key,
                    row[1].value,  # name

                    row[23].value,
                    row[24].value,
                    row[25].value
                ))
                data_count += 1
            conn.commit()
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    meta_key = 'A2385'
    id = '{}_반려동물 영화 및 공연_20200721'.format(meta_key)
    insert_excel_to_db(id, meta_key)
    # insert_excel_to_db()
