import csv
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def get_csv():
    wb = load_workbook(r'C:\Users\work\Downloads\이준재\A0111_반려동물_등록기관_20200424.xlsx', data_only=True)
    ws = wb['Sheet']
    iter_rows = iter(ws.rows)
    next(iter_rows)
    arr = []
    for row in iter_rows:
        __dict = {
            'a': row[0].value,  # ID
            'b': row[1].value,  # 데이터 분류명
            'c': row[2].value,  # 기관/회사명
            'd': row[3].value,  # 주요내용 및 서비스
            'e': row[4].value,  # 메모 및 기타
            'f': row[5].value,  # 기준년월일
            'g': row[6].value,  # 경과기간(개월)
            'h': row[7].value,  # 출처
            'i': row[8].value,  # 데이터 생성자
            'j': row[9].value,  # 데이터 저작권
            'k': row[10].value,  # 데이터 수집방벙
            'l': row[11].value,  # 데이터 프로토콜
            'm': row[12].value,  # 데이터 소스
            'n': row[13].value,  # 데이터 저장소
            'o': row[14].value,  # 데이터 확장 방법
            'p': row[15].value,  # 데이터 확장 참조
            'q': row[16].value,  # 등록자
            'r': row[17].value,  # 수정자
            's': row[18].value,  # 등록일
            't': row[19].value,  # 수정일
            'u': row[20].value,  # 키워드 -- 공통 --
            # 'v': row[21].value,  # 평점
            # 'w': row[22].value,  # 운영시간
            # 'x': row[23].value,  # 지번주소
            # 'y': row[24].value,  # 응급실 유무
        }
        arr.append(__dict)
    return arr


def save(arr):
    now = datetime.datetime.now()
    date_now = datetime.datetime.strftime(now, "%Y-%m-%d")
    file_count = 1
    wb = Workbook()
    ws = wb.active
    # 첫행 입력  a      b               c               d                   e           f               g
    ws.append(('ID', '데이터 분류명', '데이터 구분', '주요내용 및 서비스', '기준년월일', '경과기간(개월)', '출처',
               # h                  i               j               k               l               m
               '데이터 생성자', '데이터 저작권', '데이터 수집방법', '데이터 프로토콜', '데이터 소스', '데이터 저장소',
               #    n                   o               p       q        r         s        t
               '데이터 확장 방법', '데이터 확장 참조', '등록자', '수정자', '등록일', '수정일', '키워드',
               # u        v        w       x       y                 z  aa
               '주소', '전화번호'))
    # DB 모든 데이터 엑셀로
    cnt = 0

    for row in arr:
        phone = row['e'].split(',')
        try:
            cnt += 1
            a = row['a']  # ID
            b = row['b']  # 데이터 분류명
            c = '업체명'  # 데이터 구분
            d = ''  # 주요내용 및 서비스
            e = row['f']  # 기준년월일
            f = row['g']  # 경과기간
            g = row['h']  # 출처
            h = row['i']  # 데이터 생성자
            i = row['j']  # 데이터 저작권
            j = row['k']  # 데이터 수집방법
            k = row['l']  # 데이터 프로토콜
            l = row['m']  # 데이터 소스
            m = row['n']  # 데이터 저장소
            n = row['o']  # 데이터 확장 방법
            o = row['p']  # 데이터 확장 참조
            p = row['q']  # 등록자
            q = row['r']  # 수정자
            r = row['s']  # 등록일
            s = row['t']  # 수정일
            t = row['u']  # 키워드 -- 공통 --
            u = row['c'] + ' ' + row['d']  # 도로명 주소 || row['c'] + ' ' + row['d']
            v = phone[1].strip() #row['e']  # row['x']  # 전화번호 phone[1].strip()
            #w = '' # row['v']  # 평점
            #x = '' # row['x']  # 지번 주소
            #y = '' # row['w']  # 운영시간
            #z = ''
            #aa = ''
        except:
            continue
        insert = (a, b, c, d, e, f, g, h, i, j, k, l, m, n, o, p, q, r, s, t, u, v)#, w, x)#, y, z, aa)
        ws.append(insert)
        print(insert)
        # row[1] : 업체명, row[2] : 대표명, row[3] : 전화번호, row[4] : 주소
    wb.save('C:/workSpace/flask_crawling/originalDatas/A0111_반려동물_등록기관_20200722.xlsx')
# 내주변 애견카페
# 24시간 영업 애견카페
# 애견카페 평가/평점 랭킹
# 애견카페 입장료 유무


if __name__ == '__main__':
    return_arr = get_csv()
    save(return_arr)
